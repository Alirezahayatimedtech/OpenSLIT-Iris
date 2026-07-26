from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from PIL import Image

from openslit.collaboration.pilot import PilotConfig, build_pilot, sha256_file
from openslit.collaboration.profiler import build_image_profile
from openslit.collaboration.schema import (
    ALLOWED_VALUES,
    FORBIDDEN_SHARED_COLUMNS,
    REQUIRED_RESPONSE_COLUMNS,
)
from openslit.collaboration.validation import merge_submissions, validate_submission


class CollaborationPilotTest(unittest.TestCase):
    def make_fixture(self, root: Path, participants: int = 12) -> tuple[Path, Path]:
        images = root / "images"
        images.mkdir()
        rows = []
        for participant in range(participants):
            for image_number in range(2):
                path = images / f"source_{participant}_{image_number}.jpg"
                value = 20 + participant * 10 + image_number
                Image.new("RGB", (32, 32), (value, value, value)).save(path)
                rows.append(
                    {
                        "participant_id": str(participant),
                        "image_path": str(path),
                        "sha256": sha256_file(path),
                        "readable": True,
                        "brightness_mean": float(value),
                        "overexposed_fraction": float(value > 120) / 10,
                        "channel_clip_fraction": float(value > 100) / 10,
                        "laplacian_variance": float(200 - value),
                    }
                )
        profile = root / "profile.csv"
        pd.DataFrame(rows).to_csv(profile, index=False)
        manifest = root / "manifest.csv"
        pd.DataFrame(
            {
                "participant_id": [row["participant_id"] for row in rows],
                "image_path": [row["image_path"] for row in rows],
            }
        ).to_csv(manifest, index=False)
        return profile, manifest

    def test_build_is_reproducible_and_blinded(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            profile, manifest = self.make_fixture(root)
            first = root / "run1"
            second = root / "run2"
            base = dict(
                image_profile=profile,
                source_manifest=manifest,
                seed=42,
                distribution_participants=5,
                challenge_participants=5,
                double_mask_images=4,
                grader_ids=("grader_a", "grader_b"),
                copy_images=True,
            )
            build_pilot(PilotConfig(output_dir=first, **base))
            build_pilot(PilotConfig(output_dir=second, **base))
            key1 = pd.read_csv(first / "private/pilot_private_key.csv")
            key2 = pd.read_csv(second / "private/pilot_private_key.csv")
            pd.testing.assert_frame_equal(key1, key2)
            self.assertEqual(key1["participant_id"].nunique(), 10)
            self.assertEqual(key1["sha256"].nunique(), 10)
            self.assertEqual(key1["double_mask_annotation"].sum(), 4)

            shared = pd.read_csv(first / "shared/pilot_image_index.csv")
            self.assertFalse(FORBIDDEN_SHARED_COLUMNS.intersection(shared.columns))
            self.assertEqual(len(list((first / "drive_upload/images").glob("*.jpg"))), 10)

    def test_profile_uses_only_trusted_manifest_fields(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            image = root / "input.jpg"
            Image.new("RGB", (32, 32), (60, 70, 80)).save(image)
            manifest = root / "manifest.csv"
            pd.DataFrame(
                [
                    {
                        "subject": "S-1",
                        "file": image.name,
                        "laterality": "untrusted",
                        "view_label": "untrusted",
                    }
                ]
            ).to_csv(manifest, index=False)
            output = root / "profile.csv"
            profile = build_image_profile(
                manifest,
                output,
                participant_column="subject",
                image_column="file",
            )
            self.assertTrue(profile.loc[0, "readable"])
            self.assertEqual(profile.loc[0, "participant_id"], "S-1")
            self.assertNotIn("laterality", profile.columns)
            self.assertNotIn("view_label", profile.columns)

    def test_blank_workbook_is_valid_when_incomplete_allowed(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            profile, manifest = self.make_fixture(root)
            output = root / "run"
            build_pilot(
                PilotConfig(
                    image_profile=profile,
                    source_manifest=manifest,
                    output_dir=output,
                    seed=7,
                    distribution_participants=5,
                    challenge_participants=5,
                    double_mask_images=2,
                    grader_ids=("grader_a", "grader_b"),
                )
            )
            submission = output / "shared/grader_a_quality_grading.xlsx"
            index = output / "shared/pilot_image_index.csv"
            _, errors = validate_submission(submission, index, require_complete=False)
            self.assertEqual(errors, [])

            workbook = load_workbook(submission)
            self.assertEqual(
                workbook.sheetnames,
                [
                    "START HERE",
                    "Review Images",
                    "Detailed Instructions",
                    "Definitions",
                ],
            )
            self.assertEqual(workbook.active.title, "START HERE")
            sheet = workbook["Review Images"]
            headers = [cell.value for cell in sheet[1]]
            self.assertTrue(sheet.row_dimensions[1].hidden)
            self.assertEqual(sheet["E2"].value, "OPEN IMAGE")
            self.assertTrue(sheet.column_dimensions["J"].hidden)
            self.assertEqual(workbook["START HERE"]["B10"].value, 10)
            self.assertTrue(
                workbook["START HERE"]["B8"].value.startswith("=COUNTIFS(")
            )
            for row_number in range(3, sheet.max_row + 1):
                sheet.cell(
                    row=row_number,
                    column=headers.index("review_date_yyyy_mm_dd") + 1,
                    value="2026-07-26",
                )
                for field in REQUIRED_RESPONSE_COLUMNS:
                    if field not in ALLOWED_VALUES:
                        continue
                    sheet.cell(
                        row=row_number,
                        column=headers.index(field) + 1,
                        value=ALLOWED_VALUES[field][0],
                    )
            workbook.save(submission)
            _, errors = validate_submission(submission, index, require_complete=True)
            self.assertEqual(errors, [])

    def test_run_manifest_records_prohibitions(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            profile, manifest = self.make_fixture(root)
            output = root / "run"
            build_pilot(
                PilotConfig(
                    image_profile=profile,
                    source_manifest=manifest,
                    output_dir=output,
                    distribution_participants=5,
                    challenge_participants=5,
                    double_mask_images=2,
                )
            )
            run = json.loads((output / "run_manifest.json").read_text())
            self.assertFalse(run["historical_view_labels_used"])
            self.assertFalse(run["historical_laterality_used"])
            self.assertIn("laterality", run["forbidden_fields"])

    def test_completed_independent_submissions_merge(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            profile, manifest = self.make_fixture(root)
            output = root / "run"
            build_pilot(
                PilotConfig(
                    image_profile=profile,
                    source_manifest=manifest,
                    output_dir=output,
                    seed=9,
                    distribution_participants=5,
                    challenge_participants=5,
                    double_mask_images=2,
                    grader_ids=("grader_a", "grader_b"),
                )
            )
            workbooks = [
                output / "shared/grader_a_quality_grading.xlsx",
                output / "shared/grader_b_quality_grading.xlsx",
            ]
            for workbook_path in workbooks:
                workbook = load_workbook(workbook_path)
                sheet = workbook["Review Images"]
                headers = [cell.value for cell in sheet[1]]
                for row_number in range(3, sheet.max_row + 1):
                    sheet.cell(
                        row=row_number,
                        column=headers.index("review_date_yyyy_mm_dd") + 1,
                        value="2026-07-26",
                    )
                    for field in REQUIRED_RESPONSE_COLUMNS:
                        if field not in ALLOWED_VALUES:
                            continue
                        sheet.cell(
                            row=row_number,
                            column=headers.index(field) + 1,
                            value=ALLOWED_VALUES[field][0],
                        )
                workbook.save(workbook_path)
            metrics = merge_submissions(
                workbooks[0],
                workbooks[1],
                output / "shared/pilot_image_index.csv",
                output / "agreement",
            )
            self.assertEqual(metrics["requires_adjudication"], 0)
            self.assertEqual(metrics["quality_grade_quadratic_weighted_kappa"], 1.0)
            self.assertTrue((output / "agreement/adjudication_queue.csv").is_file())


if __name__ == "__main__":
    unittest.main()
