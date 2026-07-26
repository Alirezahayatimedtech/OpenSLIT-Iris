"""Create and read locked-reference collaborative grading workbooks."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .schema import (
    ALLOWED_VALUES,
    FIELD_VALUE_DEFINITIONS,
    MASK_CLASS_DEFINITIONS,
    QUALITY_GRADE_DEFINITIONS,
    REFERENCE_COLUMNS,
    RESPONSE_COLUMNS,
)

GRADING_SHEET_NAMES = ("Review Images", "quality_grading")
OPTIONAL_DETAIL_COLUMNS = {
    "focus_problem",
    "exposure_problem",
    "reflection_burden",
    "slit_beam_burden",
    "eyelid_eyelash_burden",
    "off_axis_problem",
}
FRIENDLY_HEADERS = {
    "blinded_image_id": "Image ID",
    "blinded_patient_id": "Patient alias",
    "image_file": "Image file",
    "drive_url": "Image URL",
    "open_image": "OPEN IMAGE",
    "grader_id": "Grader",
    "review_date_yyyy_mm_dd": "Review date\nYYYY-MM-DD",
    "acquisition_eligible": "Eligible?",
    "quality_grade": "Quality\nA–D",
    "focus_problem": "Focus problem",
    "exposure_problem": "Exposure problem",
    "reflection_burden": "Reflection",
    "slit_beam_burden": "Slit beam",
    "eyelid_eyelash_burden": "Lid/lashes",
    "off_axis_problem": "Off-axis",
    "pupil_visibility": "Pupil visible?",
    "outer_iris_visibility": "Outer iris visible?",
    "segmentation_feasibility": "Can it be segmented?",
    "include_for_mask_annotation": "Send to masking?",
    "primary_exclusion_reason": "Main exclusion reason",
    "confidence": "Confidence",
    "comments": "Comments (optional)",
}


def _grading_sheet(workbook):
    for name in GRADING_SHEET_NAMES:
        if name in workbook.sheetnames:
            return workbook[name]
    raise ValueError(
        f"Workbook lacks a grading sheet; expected one of {GRADING_SHEET_NAMES}"
    )


def _first_data_row(sheet) -> int:
    return 3 if sheet.cell(row=2, column=1).value == "Image ID" else 2


def write_grader_workbook(
    shared: pd.DataFrame, destination: Path, grader_id: str
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Review Images"
    columns = REFERENCE_COLUMNS + ["open_image"] + RESPONSE_COLUMNS
    sheet.append(columns)
    sheet.append([FRIENDLY_HEADERS[column] for column in columns])

    for _, source in shared.iterrows():
        row = {column: "" for column in columns}
        for column in REFERENCE_COLUMNS:
            row[column] = source.get(column, "")
        row["grader_id"] = grader_id
        sheet.append([row[column] for column in columns])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    reference_fill = PatternFill("solid", fgColor="D9EAF7")
    response_fill = PatternFill("solid", fgColor="FFF2CC")
    for header_row in (1, 2):
        for cell in sheet[header_row]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )
            cell.protection = Protection(locked=True)

    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            column = columns[cell.column - 1]
            if (
                column in REFERENCE_COLUMNS
                or column in {"open_image", "grader_id"}
            ):
                cell.fill = reference_fill
                cell.protection = Protection(locked=True)
            else:
                cell.fill = response_fill
                cell.protection = Protection(locked=False)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for column, allowed in ALLOWED_VALUES.items():
        column_index = columns.index(column) + 1
        values = ",".join(allowed)
        validation = DataValidation(
            type="list",
            formula1=f'"{values}"',
            allow_blank=True,
            error="Select one value from the controlled list.",
            errorTitle="Invalid value",
        )
        sheet.add_data_validation(validation)
        validation.add(
            f"{get_column_letter(column_index)}3:"
            f"{get_column_letter(column_index)}{len(shared) + 2}"
        )

    drive_column = columns.index("drive_url") + 1
    open_column = columns.index("open_image") + 1
    for row_number in range(3, len(shared) + 3):
        url_cell = sheet.cell(row=row_number, column=drive_column)
        if url_cell.value:
            url_cell.hyperlink = str(url_cell.value)
            url_cell.style = "Hyperlink"
        open_cell = sheet.cell(row=row_number, column=open_column)
        open_cell.value = (
            f'=IF(D{row_number}="","",HYPERLINK(D{row_number},"OPEN IMAGE"))'
        )
        open_cell.font = Font(color="0563C1", bold=True, underline="single")
        open_cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = (
        f"A2:{get_column_letter(len(columns))}{len(shared) + 2}"
    )
    sheet.row_dimensions[1].hidden = True
    sheet.row_dimensions[2].height = 42
    widths = {
        "blinded_image_id": 18,
        "blinded_patient_id": 20,
        "image_file": 24,
        "drive_url": 38,
        "open_image": 16,
        "comments": 45,
        "primary_exclusion_reason": 25,
    }
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(
            column, max(15, min(24, len(column) + 2))
        )
        if column in {
            "blinded_patient_id",
            "image_file",
            "drive_url",
            "grader_id",
        } | OPTIONAL_DETAIL_COLUMNS:
            sheet.column_dimensions[get_column_letter(index)].hidden = True
    sheet.protection.sheet = True
    sheet.protection.password = "openslit-reference"
    sheet.protection.autoFilter = False
    sheet.protection.sort = False

    start = workbook.create_sheet("START HERE", 0)
    start_rows = [
        ("OPENSLIT-IRIS IMAGE REVIEW", ""),
        ("Your task", "Review every image independently."),
        ("1", "Open the Review Images tab."),
        ("2", "Click OPEN IMAGE for the first row."),
        ("3", "Complete the visible yellow cells using dropdowns."),
        ("4", "Repeat until Remaining is 0, then submit this file."),
        (
            "Important",
            "Grade only what is visible. Do not infer left/right, laterality, "
            "center, nasal, or temporal.",
        ),
        (
            "Completed",
            "=COUNTIFS("
            + ",".join(
                f"'Review Images'!{column}3:{column}{len(shared) + 2},\"<>\""
                for column in ("G", "H", "I", "P", "Q", "R", "S", "T", "U")
            )
            + ")",
        ),
        ("Remaining", "=B10-B8"),
        ("Total", len(shared)),
        ("Quality grade", "Use the definitions below."),
        ("A", QUALITY_GRADE_DEFINITIONS["A"]),
        ("B", QUALITY_GRADE_DEFINITIONS["B"]),
        ("C", QUALITY_GRADE_DEFINITIONS["C"]),
        ("D", QUALITY_GRADE_DEFINITIONS["D"]),
        (
            "Optional detail",
            "Focus, exposure, reflection, slit-beam, lid/lash, and off-axis "
            "fields are hidden. The custodian can unhide them when needed.",
        ),
    ]
    for row in start_rows:
        start.append(row)
    start.merge_cells("A1:B1")
    start["A1"].fill = header_fill
    start["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    start["A1"].alignment = Alignment(horizontal="center", vertical="center")
    start.row_dimensions[1].height = 32
    start.column_dimensions["A"].width = 22
    start.column_dimensions["B"].width = 100
    for row in start.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row_number in (8, 9, 10):
        for cell in start[row_number]:
            cell.fill = response_fill
            cell.font = Font(bold=True)
    start.freeze_panes = "A2"

    instructions = workbook.create_sheet("Detailed Instructions")
    instructions_rows = [
        ("Purpose", "Independent blinded quality and eligibility grading."),
        (
            "Independence",
            "Do not view another grader's responses before submitting this workbook.",
        ),
        (
            "Anatomy",
            "Grade only visible structures. Never infer hidden tissue.",
        ),
        (
            "Prohibited labels",
            "Do not assign or infer left/right, laterality, center, nasal, or temporal.",
        ),
        (
            "Required",
            "Complete the visible yellow cells. Comments and hidden detail "
            "fields are optional.",
        ),
        (
            "Quality A",
            QUALITY_GRADE_DEFINITIONS["A"],
        ),
        ("Quality B", QUALITY_GRADE_DEFINITIONS["B"]),
        ("Quality C", QUALITY_GRADE_DEFINITIONS["C"]),
        ("Quality D", QUALITY_GRADE_DEFINITIONS["D"]),
        (
            "Mask work",
            "Draw masks in CVAT or equivalent software, not inside this spreadsheet.",
        ),
    ]
    for row in instructions_rows:
        instructions.append(row)
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 105
    for cell in instructions[1]:
        cell.font = Font(bold=True)
    for row in instructions.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    codebook = workbook.create_sheet("Definitions")
    codebook.append(["field", "allowed_value", "definition"])
    for grade, definition in QUALITY_GRADE_DEFINITIONS.items():
        codebook.append(["quality_grade", grade, definition])
    for field, values in ALLOWED_VALUES.items():
        if field == "quality_grade":
            continue
        for value in values:
            codebook.append(
                [field, value, FIELD_VALUE_DEFINITIONS.get((field, value), "")]
            )
    for class_id, definition in MASK_CLASS_DEFINITIONS.items():
        codebook.append(["mask_class", str(class_id), definition])
    codebook.column_dimensions["A"].width = 32
    codebook.column_dimensions["B"].width = 28
    codebook.column_dimensions["C"].width = 85
    for cell in codebook[1]:
        cell.font = Font(bold=True)
    for row in codebook.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.active = 0
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)


def read_grading_table(path: Path) -> pd.DataFrame:
    if path.suffix.lower() == ".csv":
        return pd.read_csv(path, dtype=str, keep_default_na=False)
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = _grading_sheet(workbook)
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"{path} contains no grading rows")
    headers = [str(value) if value is not None else "" for value in rows[0]]
    first_data_index = 2 if len(rows) > 1 and rows[1][0] == "Image ID" else 1
    data = [
        ["" if value is None else str(value).strip() for value in row]
        for row in rows[first_data_index:]
    ]
    return pd.DataFrame(data, columns=headers)


def apply_drive_links(workbook_path: Path, links: pd.DataFrame) -> None:
    required = {"blinded_image_id", "drive_url"}
    if not required.issubset(links.columns):
        raise ValueError(f"Drive-link table must contain {sorted(required)}")
    link_map = dict(zip(links["blinded_image_id"], links["drive_url"]))
    workbook = load_workbook(workbook_path)
    sheet = _grading_sheet(workbook)
    headers = [cell.value for cell in sheet[1]]
    id_column = headers.index("blinded_image_id") + 1
    url_column = headers.index("drive_url") + 1
    open_column = (
        headers.index("open_image") + 1 if "open_image" in headers else None
    )
    for row_number in range(_first_data_row(sheet), sheet.max_row + 1):
        image_id = sheet.cell(row=row_number, column=id_column).value
        url = str(link_map.get(image_id, "") or "").strip()
        cell = sheet.cell(row=row_number, column=url_column)
        cell.value = url
        cell.hyperlink = url or None
        if url:
            cell.style = "Hyperlink"
        if open_column:
            open_cell = sheet.cell(row=row_number, column=open_column)
            url_letter = get_column_letter(url_column)
            open_cell.value = (
                f'=IF({url_letter}{row_number}="","",'
                f'HYPERLINK({url_letter}{row_number},"OPEN IMAGE"))'
            )
    workbook.save(workbook_path)


def apply_drive_links_to_csv(table_path: Path, links: pd.DataFrame) -> None:
    required = {"blinded_image_id", "drive_url"}
    if not required.issubset(links.columns):
        raise ValueError(f"Drive-link table must contain {sorted(required)}")
    table = pd.read_csv(table_path, dtype=str, keep_default_na=False)
    if "blinded_image_id" not in table.columns or "drive_url" not in table.columns:
        raise ValueError(f"{table_path} lacks blinded_image_id or drive_url")
    link_map = dict(zip(links["blinded_image_id"], links["drive_url"]))
    missing = set(table["blinded_image_id"]) - set(link_map)
    if missing:
        raise ValueError(f"Drive links missing image IDs: {sorted(missing)}")
    table["drive_url"] = table["blinded_image_id"].map(link_map).fillna("")
    table.to_csv(table_path, index=False)
